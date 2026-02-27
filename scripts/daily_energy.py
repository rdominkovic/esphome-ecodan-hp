# ABOUTME: Generates daily energy summary Excel file from heat pump CSV logs.
# ABOUTME: Monthly sheets with separated consumed/delivered tables. Supports historical data.

import csv
import json
import os
from datetime import datetime
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

LOG_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
LOG_FILE = os.path.join(LOG_DIR, "ecodan_log.csv")
XLSX_FILE = os.path.join(LOG_DIR, "daily_energy.xlsx")
CSV_OUT_FILE = os.path.join(LOG_DIR, "daily_energy.csv")
HISTORICAL_FILE = os.path.join(LOG_DIR, "historical_energy.json")


def load_daily_data():
    """Read CSV log and group by date, extracting energy counters."""
    days = defaultdict(list)
    with open(LOG_FILE) as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                ts = datetime.strptime(row["timestamp"], "%Y-%m-%d %H:%M:%S")
                date_key = ts.date()
                entry = {"_ts": ts}
                for col in ["heating_consumed_kwh", "heating_delivered_kwh",
                            "dhw_consumed_kwh", "dhw_delivered_kwh",
                            "daily_consumed_kwh", "daily_produced_kwh",
                            "outside_temp", "compressor_on"]:
                    try:
                        entry[col] = float(row[col])
                    except (ValueError, TypeError, KeyError):
                        entry[col] = None
                days[date_key].append(entry)
            except (ValueError, KeyError):
                continue
    return days


def calc_daily_summary(rows):
    """Calculate energy summary for one day from its rows."""
    if len(rows) < 2:
        return None

    first = rows[0]
    last = rows[-1]

    def delta(col):
        v0 = first.get(col)
        v1 = last.get(col)
        if v0 is not None and v1 is not None and v1 >= v0:
            return round(v1 - v0, 2)
        return None

    def safe_cop(delivered, consumed):
        if consumed and consumed > 0.1 and delivered is not None:
            return round(delivered / consumed, 1)
        return None

    htg_con = delta("heating_consumed_kwh")
    htg_del = delta("heating_delivered_kwh")
    dhw_con = delta("dhw_consumed_kwh")
    dhw_del = delta("dhw_delivered_kwh")

    def daily_max(col):
        """Daily counters reset at midnight; use max to handle race condition."""
        vals = [r.get(col) for r in rows if r.get(col) is not None]
        return round(max(vals), 2) if vals else None

    tot_con = daily_max("daily_consumed_kwh")
    tot_del = daily_max("daily_produced_kwh")

    if tot_con is None and htg_con is not None:
        tot_con = round((htg_con or 0) + (dhw_con or 0), 2)
    if tot_del is None and htg_del is not None:
        tot_del = round((htg_del or 0) + (dhw_del or 0), 2)

    # Individual counters update infrequently; when they show 0 delta
    # but daily counters have data, attribute remainder to heating.
    if tot_con is not None and tot_con > 0.1:
        known_con = (htg_con or 0) + (dhw_con or 0)
        if known_con < tot_con * 0.5:
            htg_con = round(tot_con - (dhw_con or 0), 2)
    if tot_del is not None and tot_del > 0.1:
        known_del = (htg_del or 0) + (dhw_del or 0)
        if known_del < tot_del * 0.5:
            htg_del = round(tot_del - (dhw_del or 0), 2)

    out_temps = [r["outside_temp"] for r in rows if r["outside_temp"] is not None]
    avg_out = round(sum(out_temps) / len(out_temps), 1) if out_temps else None

    on_rows = [r for r in rows if r.get("compressor_on") == 1]
    on_pct = round(len(on_rows) / len(rows) * 100) if rows else None

    hours = (last["_ts"] - first["_ts"]).total_seconds() / 3600

    return {
        "date": first["_ts"].date(),
        "hours": round(hours, 1),
        "samples": len(rows),
        "avg_outside": avg_out,
        "comp_on_pct": on_pct,
        "htg_consumed": htg_con,
        "htg_delivered": htg_del,
        "htg_cop": safe_cop(htg_del, htg_con),
        "dhw_consumed": dhw_con,
        "dhw_delivered": dhw_del,
        "dhw_cop": safe_cop(dhw_del, dhw_con),
        "total_consumed": tot_con,
        "total_delivered": tot_del,
        "total_cop": safe_cop(tot_del, tot_con),
    }


def load_historical():
    """Load historical monthly energy totals from JSON file."""
    if not os.path.exists(HISTORICAL_FILE):
        return {}
    with open(HISTORICAL_FILE) as f:
        return json.load(f)


CONSUMED_COLUMNS = [
    ("Date", "date", 12),
    ("Out \u00b0C", "avg_outside", 8),
    ("Comp %", "comp_on_pct", 8),
    ("Htg kWh", "htg_consumed", 10),
    ("DHW kWh", "dhw_consumed", 10),
    ("Total kWh", "total_consumed", 11),
]

DELIVERED_COLUMNS = [
    ("Date", "date", 12),
    ("Out \u00b0C", "avg_outside", 8),
    ("Htg kWh", "htg_delivered", 10),
    ("Htg COP", "htg_cop", 8),
    ("DHW kWh", "dhw_delivered", 10),
    ("DHW COP", "dhw_cop", 8),
    ("Total kWh", "total_delivered", 11),
    ("COP", "total_cop", 7),
]


def _make_styles():
    return {
        "header_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "header_font": Font(bold=True, size=11, color="FFFFFF"),
        "title_font": Font(bold=True, size=12),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
        "total_fill": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        "total_font": Font(bold=True, size=11),
    }


def _write_table(ws, start_row, title, columns, summaries, styles):
    """Write a formatted table to worksheet. Returns next available row."""
    border = styles["border"]

    # Title
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font = styles["title_font"]
    start_row += 1

    # Headers
    for ci, (label, _, width) in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=ci, value=label)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        cur = ws.column_dimensions[get_column_letter(ci)].width or 0
        ws.column_dimensions[get_column_letter(ci)].width = max(cur, width)
    start_row += 1

    # Data rows
    for s in summaries:
        for ci, (_, key, _) in enumerate(columns, 1):
            val = s.get(key)
            if key == "date":
                val = val.strftime("%Y-%m-%d") if val else ""
            cell = ws.cell(row=start_row, column=ci, value=val)
            cell.border = border
            if key == "date":
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                if key.endswith("_cop") and val is not None:
                    cell.number_format = "0.0"
                elif isinstance(val, float):
                    cell.number_format = "0.00"
        start_row += 1

    # Totals row
    for ci, (_, key, _) in enumerate(columns, 1):
        if key == "date":
            val = "Total"
            cell = ws.cell(row=start_row, column=ci, value=val)
            cell.alignment = Alignment(horizontal="left")
        elif key in ("avg_outside", "comp_on_pct"):
            vals = [s[key] for s in summaries if s.get(key) is not None]
            val = round(sum(vals) / len(vals), 1) if vals else None
            cell = ws.cell(row=start_row, column=ci, value=val)
            cell.number_format = "0.0"
            cell.alignment = Alignment(horizontal="right")
        elif key.endswith("_cop"):
            # Weighted COP = sum(delivered) / sum(consumed)
            del_key = key.replace("_cop", "_delivered")
            con_key = key.replace("_cop", "_consumed")
            del_sum = sum(s.get(del_key, 0) or 0 for s in summaries)
            con_sum = sum(s.get(con_key, 0) or 0 for s in summaries)
            val = round(del_sum / con_sum, 1) if con_sum > 0.1 else None
            cell = ws.cell(row=start_row, column=ci, value=val)
            cell.number_format = "0.0"
            cell.alignment = Alignment(horizontal="right")
        else:
            vals = [s.get(key, 0) or 0 for s in summaries]
            val = round(sum(vals), 2) if any(s.get(key) is not None for s in summaries) else None
            cell = ws.cell(row=start_row, column=ci, value=val)
            if val is not None:
                cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="right")
        cell.fill = styles["total_fill"]
        cell.font = styles["total_font"]
        cell.border = border

    return start_row + 2


def _aggregate_month(month_key, by_month, historical):
    """Combine historical and logged data for a month into summary dict."""
    hist = historical.get(month_key, {})
    daily = by_month.get(month_key, [])

    def _sum_daily(key):
        return sum(s.get(key, 0) or 0 for s in daily)

    # Daily counter may report total but not htg/dhw breakdown.
    # Attribute unbroken remainder to heating (primary use in winter).
    daily_htg_con = _sum_daily("htg_consumed")
    daily_dhw_con = _sum_daily("dhw_consumed")
    daily_total_con = _sum_daily("total_consumed")
    daily_htg_del = _sum_daily("htg_delivered")
    daily_dhw_del = _sum_daily("dhw_delivered")
    daily_total_del = _sum_daily("total_delivered")

    unattrib_con = max(0, daily_total_con - daily_htg_con - daily_dhw_con)
    unattrib_del = max(0, daily_total_del - daily_htg_del - daily_dhw_del)

    htg_con = (hist.get("htg_consumed_kwh") or 0) + daily_htg_con + unattrib_con
    dhw_con = (hist.get("dhw_consumed_kwh") or 0) + daily_dhw_con
    htg_del = (hist.get("htg_delivered_kwh") or 0) + daily_htg_del + unattrib_del
    dhw_del = (hist.get("dhw_delivered_kwh") or 0) + daily_dhw_del

    return {
        "month": month_key,
        "htg_consumed": round(htg_con, 2),
        "dhw_consumed": round(dhw_con, 2),
        "total_consumed": round(htg_con + dhw_con, 2),
        "htg_delivered": round(htg_del, 2),
        "dhw_delivered": round(dhw_del, 2),
        "total_delivered": round(htg_del + dhw_del, 2),
    }


def _write_summary_table(ws, start_row, title, columns, rows, styles):
    """Write a summary table. columns: list of (label, key, width). Returns next row."""
    border = styles["border"]

    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font = styles["title_font"]
    start_row += 1

    for ci, (label, _, width) in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=ci, value=label)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        cur = ws.column_dimensions[get_column_letter(ci)].width or 0
        ws.column_dimensions[get_column_letter(ci)].width = max(cur, width)
    start_row += 1

    for r in rows:
        for ci, (_, key, _) in enumerate(columns, 1):
            val = r.get(key)
            cell = ws.cell(row=start_row, column=ci, value=val if val else None)
            cell.border = border
            if key != "month":
                cell.alignment = Alignment(horizontal="right")
                if key == "cop":
                    cell.number_format = "0.0"
                elif isinstance(val, float):
                    cell.number_format = "0.00"
        start_row += 1

    # Yearly total
    totals = {}
    for _, key, _ in columns:
        if key == "month":
            totals[key] = datetime.now().strftime("%Y") + " Total"
        elif key == "cop":
            del_sum = sum(r.get("total_delivered", 0) or 0 for r in rows)
            con_sum = sum(r.get("total_consumed", 0) or 0 for r in rows)
            totals[key] = round(del_sum / con_sum, 1) if con_sum > 0.1 else None
        else:
            totals[key] = round(sum(r.get(key, 0) or 0 for r in rows), 2)

    for ci, (_, key, _) in enumerate(columns, 1):
        val = totals.get(key)
        cell = ws.cell(row=start_row, column=ci, value=val if val else None)
        cell.fill = styles["total_fill"]
        cell.font = styles["total_font"]
        cell.border = border
        if key != "month":
            cell.alignment = Alignment(horizontal="right")
            if key == "cop":
                cell.number_format = "0.0"
            elif isinstance(val, float):
                cell.number_format = "0.00"

    return start_row + 2


SUMMARY_CONSUMED = [
    ("Month", "month", 12),
    ("Htg kWh", "htg_consumed", 10),
    ("DHW kWh", "dhw_consumed", 10),
    ("Total kWh", "total_consumed", 11),
]

SUMMARY_DELIVERED = [
    ("Month", "month", 12),
    ("Htg kWh", "htg_delivered", 10),
    ("DHW kWh", "dhw_delivered", 10),
    ("Total kWh", "total_delivered", 11),
    ("COP", "cop", 7),
]


def _write_summary_sheet(wb, by_month, historical, styles):
    """Write summary sheet with monthly consumed/delivered tables and yearly totals."""
    ws = wb.create_sheet(title="Summary")

    all_months = sorted(set(list(historical.keys()) + list(by_month.keys())))
    month_rows = []
    for month_key in all_months:
        agg = _aggregate_month(month_key, by_month, historical)
        total_con = agg["total_consumed"]
        total_del = agg["total_delivered"]
        agg["cop"] = round(total_del / total_con, 1) if total_con > 0.1 else None
        month_rows.append(agg)

    row = 1
    row = _write_summary_table(ws, row, "Consumed Energy (kWh)", SUMMARY_CONSUMED, month_rows, styles)
    row = _write_summary_table(ws, row, "Delivered Heat (kWh)", SUMMARY_DELIVERED, month_rows, styles)


def write_xlsx(summaries):
    """Write daily summaries to Excel with monthly sheets and separated tables."""
    wb = Workbook()
    styles = _make_styles()

    by_month = defaultdict(list)
    for s in summaries:
        by_month[s["date"].strftime("%Y-%m")].append(s)

    historical = load_historical()

    first_sheet = True
    for month_key in sorted(by_month.keys()):
        month_data = by_month[month_key]
        if first_sheet:
            ws = wb.active
            ws.title = month_key
            first_sheet = False
        else:
            ws = wb.create_sheet(title=month_key)

        row = 1
        row = _write_table(ws, row, "Consumed Energy (kWh)", CONSUMED_COLUMNS, month_data, styles)
        row = _write_table(ws, row, "Delivered Heat (kWh)", DELIVERED_COLUMNS, month_data, styles)

    _write_summary_sheet(wb, by_month, historical, styles)

    if first_sheet:
        wb.active.title = "Summary"

    wb.save(XLSX_FILE)
    print(f"Written {XLSX_FILE} ({len(summaries)} days, {len(by_month)} months)")


def write_csv_fallback(summaries):
    """Write daily summaries to CSV if openpyxl is unavailable."""
    all_cols = [
        ("Date", "date"), ("Out \u00b0C", "avg_outside"), ("Comp %", "comp_on_pct"),
        ("Htg Con kWh", "htg_consumed"), ("Htg Del kWh", "htg_delivered"), ("Htg COP", "htg_cop"),
        ("DHW Con kWh", "dhw_consumed"), ("DHW Del kWh", "dhw_delivered"), ("DHW COP", "dhw_cop"),
        ("Total Con kWh", "total_consumed"), ("Total Del kWh", "total_delivered"), ("Total COP", "total_cop"),
    ]
    with open(CSV_OUT_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([label for label, _ in all_cols])
        for s in summaries:
            row = []
            for _, key in all_cols:
                val = s.get(key)
                if key == "date":
                    val = val.strftime("%Y-%m-%d") if val else ""
                row.append(val if val is not None else "")
            writer.writerow(row)
    print(f"Written {CSV_OUT_FILE} ({len(summaries)} days)")


def main():
    days = load_daily_data()
    summaries = []
    for date_key in sorted(days.keys()):
        s = calc_daily_summary(days[date_key])
        if s:
            summaries.append(s)

    if not summaries:
        print("No data found.")
        return

    # Console summary
    print()
    print(f"{'Date':<12} {'Out':>5} {'Htg In':>7} {'Htg Out':>8} {'hCOP':>5} "
          f"{'DHW In':>7} {'DHW Out':>8} {'dCOP':>5} "
          f"{'Tot In':>7} {'Tot Out':>8} {'COP':>5}")
    print("-" * 95)
    for s in summaries:
        def f(v, fmt=".2f"):
            return f"{v:{fmt}}" if v is not None else "-"
        print(f"{s['date']!s:<12} {f(s['avg_outside'], '.0f'):>5} "
              f"{f(s['htg_consumed']):>7} {f(s['htg_delivered']):>8} {f(s['htg_cop'], '.1f'):>5} "
              f"{f(s['dhw_consumed']):>7} {f(s['dhw_delivered']):>8} {f(s['dhw_cop'], '.1f'):>5} "
              f"{f(s['total_consumed']):>7} {f(s['total_delivered']):>8} {f(s['total_cop'], '.1f'):>5}")
    print()

    if HAS_OPENPYXL:
        write_xlsx(summaries)
    else:
        print("openpyxl not installed, writing CSV instead.")
        write_csv_fallback(summaries)


if __name__ == "__main__":
    main()
